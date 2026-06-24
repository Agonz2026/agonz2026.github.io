from openpyxl import load_workbook
from openpyxl.styles import Protection
F="PI_Case_Valuation_Calculator_FL.xlsx"
wb=load_workbook(F)
ws=wb["Calculator"]
# every cell a case manager is allowed to type in
inputs=["C5","C6","C7","C8","C11","C12","C13","C14","C15","C20","C22","D22",
 "C27","D27","C28","C36","C37","C38","C39","C40","C50","C52","C54","C62"]
unlock=Protection(locked=False)
for a in inputs: ws[a].protection=unlock
ws.protection.sheet=True
ws.protection.password="pi"          # firm can unprotect to edit formulas
ws.protection.selectLockedCells=False   # allow click/copy on locked cells
ws.protection.selectUnlockedCells=False
ws.protection.formatColumns=True; ws.protection.formatRows=True
wb.save(F)
print("locked")
