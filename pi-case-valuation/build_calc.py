from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

wb = Workbook()
ARIAL = "Arial"
title_f=Font(name=ARIAL,size=16,bold=True,color="FFFFFF")
sub_f=Font(name=ARIAL,size=9,italic=True,color="FFFFFF")
sect_f=Font(name=ARIAL,size=11,bold=True,color="FFFFFF")
lbl_f=Font(name=ARIAL,size=10); lbl_b=Font(name=ARIAL,size=10,bold=True)
input_f=Font(name=ARIAL,size=10,color="0000FF"); calc_f=Font(name=ARIAL,size=10,color="000000")
calc_b=Font(name=ARIAL,size=11,bold=True,color="000000"); note_f=Font(name=ARIAL,size=8,italic=True,color="808080")
result_f=Font(name=ARIAL,size=12,bold=True,color="000000")
navy=PatternFill("solid",fgColor="1F3864"); blue2=PatternFill("solid",fgColor="2E5496")
yellow=PatternFill("solid",fgColor="FFF2CC"); inputfill=PatternFill("solid",fgColor="FFFDE7")
green=PatternFill("solid",fgColor="E2EFDA"); result_fill=PatternFill("solid",fgColor="FCE4D6")
thin=Side(style="thin",color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
CUR='$#,##0;($#,##0);"-"'; MULT='0.0"x"'; PCT='0.0%'

# ---------------- REFERENCE ----------------
ref=wb.active; ref.title="Reference"
ref["A1"]="Severity"; ref["B1"]="Mult Low"; ref["C1"]="Mult High"; ref["D1"]="Description"
sev=[("1 - Minor (soft tissue, full recovery)",1.5,2.5,"Sprains/strains, conservative care, no residuals"),
 ("2 - Moderate (treatment, minor residual)",2.5,3.5,"Injections, PT, some lasting discomfort"),
 ("3 - Serious (surgery / objective injury)",3.5,4.5,"Fracture, surgery, MRI-confirmed injury"),
 ("4 - Severe (permanent / disfigurement)",4.5,6.0,"Permanent impairment, scarring, chronic pain"),
 ("5 - Catastrophic (disability, brain/spinal)",5.0,10.0,"TBI, paralysis, amputation, lifelong care")]
for i,r in enumerate(sev,2):
    for j,v in enumerate(r,1): ref.cell(i,j,v)
ref["F1"]="Liability"; ref["G1"]="Factor"; ref["H1"]="Meaning"
liab=[("Clear / strong",1.00,"Liability admitted or near-certain"),
 ("Probable",0.85,"Good facts, minor disputes"),
 ("Disputed",0.65,"Genuine fact dispute on fault"),
 ("Weak / questionable",0.40,"Liability is a real problem")]
for i,r in enumerate(liab,2):
    ref.cell(i,6,r[0]); ref.cell(i,7,r[1]); ref.cell(i,8,r[2])
ref["J1"]="Age from"; ref["K1"]="Factor"; ref["L1"]="Bracket"
age=[(0,1.15,"Under 18"),(18,1.15,"18-29"),(30,1.05,"30-49"),(50,1.00,"50-64"),(65,0.92,"65-79"),(80,0.85,"80+")]
for i,r in enumerate(age,2):
    ref.cell(i,10,r[0]); ref.cell(i,11,r[1]); ref.cell(i,12,r[2])
ref["N1"]="Prior history"; ref["O1"]="Factor"
prior=[("None / clean history",1.00),("Prior - unrelated body part",0.97),
 ("Prior - same body part (documented)",0.80),("Prior - same part + recent treatment",0.70)]
for i,r in enumerate(prior,2):
    ref.cell(i,14,r[0]); ref.cell(i,15,r[1])
ref["Q1"]="Vehicle / impact severity"; ref["R1"]="Factor"
pd=[("Severe - totaled / major impact",1.10),("Moderate - drivable with damage",1.00),
 ("Minor - cosmetic / low impact",0.85)]
for i,r in enumerate(pd,2):
    ref.cell(i,17,r[0]); ref.cell(i,18,r[1])
ref["A8"]="Impairment factor: 1 + (whole-person impairment % x 1.5), capped at 1.6"
ref["A9"]="Property damage is NOT a recoverable element of the BI claim - used only as a causation/severity signal."
ref["A8"].font=note_f; ref["A9"].font=note_f
for c in ["A1","B1","C1","D1","F1","G1","H1","J1","K1","L1","N1","O1","Q1","R1"]: ref[c].font=lbl_b
ref.column_dimensions["A"].width=42; ref.column_dimensions["D"].width=42
ref.column_dimensions["F"].width=20; ref.column_dimensions["H"].width=38
ref.column_dimensions["L"].width=12; ref.column_dimensions["N"].width=34; ref.column_dimensions["Q"].width=32

# ---------------- CALCULATOR ----------------
ws=wb.create_sheet("Calculator",0); ws.sheet_view.showGridLines=False
ws.column_dimensions["A"].width=2; ws.column_dimensions["B"].width=46
ws.column_dimensions["C"].width=18; ws.column_dimensions["D"].width=18; ws.column_dimensions["E"].width=46
def section(r,t):
    ws.merge_cells(f"B{r}:E{r}"); c=ws[f"B{r}"]; c.value=t; c.font=sect_f; c.fill=blue2
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1); ws.row_dimensions[r].height=20
def label(r,t,bold=False,note=None):
    c=ws[f"B{r}"]; c.value=t; c.font=lbl_b if bold else lbl_f; c.alignment=Alignment(indent=1,vertical="center")
    if note: n=ws[f"E{r}"]; n.value=note; n.font=note_f; n.alignment=Alignment(indent=1,vertical="center")
def inp(r,val,fmt=CUR,col="C"):
    c=ws[f"{col}{r}"]; c.value=val; c.font=input_f; c.fill=inputfill; c.number_format=fmt
    c.border=border; c.alignment=Alignment(horizontal="right"); return c
def out(r,f,fmt=CUR,col="C",bold=False,fill=None):
    c=ws[f"{col}{r}"]; c.value=f; c.font=calc_b if bold else calc_f; c.number_format=fmt
    c.border=border; c.alignment=Alignment(horizontal="right")
    if fill: c.fill=fill
    return c

ws.merge_cells("B1:E1"); ws["B1"]="PERSONAL INJURY CASE VALUATION CALCULATOR"
ws["B1"].font=title_f; ws["B1"].fill=navy; ws["B1"].alignment=Alignment(horizontal="left",vertical="center",indent=1)
ws.row_dimensions[1].height=30
ws.merge_cells("B2:E2"); ws["B2"]="Florida  |  Hybrid multiplier + per-diem range  |  Blue cells = inputs. Edit only blue/yellow cells."
ws["B2"].font=sub_f; ws["B2"].fill=navy; ws["B2"].alignment=Alignment(horizontal="left",indent=1)

section(4,"CASE INFORMATION")
for i,lab in enumerate(["Client name","Claim / file no.","Date of loss","Adjuster / carrier"]):
    r=5+i; label(r,lab); ws.merge_cells(f"C{r}:E{r}"); c=ws[f"C{r}"]
    c.font=input_f; c.fill=inputfill; c.border=border; c.alignment=Alignment(indent=1)
ws["C9"]="LOW"; ws["D9"]="HIGH"
for cc in ("C9","D9"): ws[cc].font=Font(name=ARIAL,size=8,bold=True,color="808080"); ws[cc].alignment=Alignment(horizontal="center")

section(10,"1.  ECONOMIC DAMAGES (SPECIALS - recoverable)")
label(11,"Past medical bills (billed)"); inp(11,0)
label(12,"Future medical costs (estimated)"); inp(12,0)
label(13,"Past lost wages"); inp(13,0)
label(14,"Future lost earnings / earning capacity"); inp(14,0)
label(15,"Other out-of-pocket (mileage, meds, devices)"); inp(15,0)
label(16,"Total economic damages",bold=True); out(16,"=SUM(C11:C15)",bold=True,fill=green)
label(17,"Medical specials (past + future)  - multiplier base"); out(17,"=C11+C12")

section(19,"2.  PAIN & SUFFERING - METHOD A: MULTIPLIER")
label(20,"Injury severity (choose)")
sc=ws["C20"]; ws.merge_cells("C20:E20"); sc.value=sev[0][0]; sc.font=input_f; sc.fill=yellow
sc.border=border; sc.alignment=Alignment(indent=1)
label(21,"Suggested multiplier (from severity)")
out(21,"=VLOOKUP($C$20,Reference!$A$2:$C$6,2,FALSE)",MULT,col="C")
out(21,"=VLOOKUP($C$20,Reference!$A$2:$C$6,3,FALSE)",MULT,col="D")
ws["C21"].alignment=Alignment(horizontal="center"); ws["D21"].alignment=Alignment(horizontal="center")
ws["E21"].value="low  /  high"; ws["E21"].font=note_f; ws["E21"].alignment=Alignment(indent=1)
label(22,"Override multiplier (optional)",note="Leave blank to use suggested")
inp(22,None,MULT,col="C"); inp(22,None,MULT,col="D")
ws["C22"].alignment=Alignment(horizontal="center"); ws["D22"].alignment=Alignment(horizontal="center")
label(23,"Multiplier used")
out(23,'=IF(C22="",C21,C22)',MULT,col="C"); out(23,'=IF(D22="",D21,D22)',MULT,col="D")
ws["C23"].alignment=Alignment(horizontal="center"); ws["D23"].alignment=Alignment(horizontal="center")
label(24,"Pain & suffering - Method A (low / high)")
out(24,"=$C$17*C23",col="C"); out(24,"=$C$17*D23",col="D")

section(26,"3.  PAIN & SUFFERING - METHOD B: PER DIEM")
label(27,"Daily rate (low / high)",note="Common: $150-$300/day")
inp(27,150,CUR,col="C"); inp(27,300,CUR,col="D")
label(28,"Days of pain / impact",note="Injury date to MMI, or treatment days")
inp(28,0,'#,##0',col="C")
label(29,"Pain & suffering - Method B (low / high)")
out(29,"=C27*$C$28",col="C"); out(29,"=D27*$C$28",col="D")

section(31,"4.  BLENDED NON-ECONOMIC (BASE)")
label(32,"Non-economic base (low / high)",bold=True)
out(32,"=IF(AND(C24>0,C29>0),MIN(C24,C29),MAX(C24,C29))",col="C",bold=True,fill=green)
out(32,"=MAX(D24,D29)",col="D",bold=True,fill=green)
label(33,"Method note",note="If only one method is used, that method sets the range")

section(35,"5.  CASE CHARACTERISTICS (adjust pain & suffering)")
label(36,"Permanent impairment rating (whole-person %)",note="Treating physician / IME, AMA Guides")
inp(36,0,PCT)
out(36,"=1+MIN(C36,0.4)*1.5",MULT,col="D"); ws["D36"].alignment=Alignment(horizontal="center")
ws["E36"].value="factor"; ws["E36"].font=note_f; ws["E36"].alignment=Alignment(indent=1)
label(37,"Client age (years)",note="Younger w/ permanent injury = longer P&S")
inp(37,40,'0')
out(37,"=VLOOKUP(C37,Reference!$J$2:$K$7,2,TRUE)",MULT,col="D"); ws["D37"].alignment=Alignment(horizontal="center")
ws["E37"].value="factor"; ws["E37"].font=note_f; ws["E37"].alignment=Alignment(indent=1)
label(38,"Prior accidents / injuries (choose)")
pc=ws["C38"]; pc.value=prior[0][0]; pc.font=input_f; pc.fill=yellow; pc.border=border; pc.alignment=Alignment(indent=1)
out(38,"=VLOOKUP(C38,Reference!$N$2:$O$5,2,FALSE)",MULT,col="D"); ws["D38"].alignment=Alignment(horizontal="center")
ws["E38"].value="factor"; ws["E38"].font=note_f; ws["E38"].alignment=Alignment(indent=1)
label(39,"Property damage $ (NOT recoverable - signal only)",note="Used for causation check, not added to value")
inp(39,0)
label(40,"Vehicle / impact severity (choose)")
vc=ws["C40"]; vc.value=pd[1][0]; vc.font=input_f; vc.fill=yellow; vc.border=border; vc.alignment=Alignment(indent=1)
out(40,"=VLOOKUP(C40,Reference!$Q$2:$R$4,2,FALSE)",MULT,col="D"); ws["D40"].alignment=Alignment(horizontal="center")
ws["E40"].value="factor"; ws["E40"].font=note_f; ws["E40"].alignment=Alignment(indent=1)
label(41,"Causation check (PD vs. medicals)")
cc=ws["C41"]; ws.merge_cells("C41:E41")
cc.value=('=IF(AND($C$39>0,$C$17>0,$C$39<$C$17*0.15),'
          '"LOW property damage vs. high medicals - causation EXPOSURE",'
          'IF($C$40="Severe - totaled / major impact","High impact - causation & damages well-supported",'
          'IF($C$39=0,"Enter PD to run causation check","Property damage consistent with claimed injuries")))')
cc.font=calc_b; cc.alignment=Alignment(indent=1)
label(42,"Composite non-economic modifier",bold=True,note="Impairment x age x prior x impact (capped 0.4x-2.0x)")
out(42,"=MAX(0.4,MIN(2,D36*D37*D38*D40))",MULT,bold=True,fill=green); ws["C42"].alignment=Alignment(horizontal="center")
label(43,"Florida permanency threshold check")
tc=ws["C43"]; ws.merge_cells("C43:E43")
tc.value=('=IF(C36>0,"Impairment shown - non-economic threshold likely met",'
          '"No impairment % - confirm permanency threshold (Fla. Stat. 627.737)")')
tc.font=calc_b; tc.alignment=Alignment(indent=1)
label(44,"Adjusted non-economic (low / high)",bold=True)
out(44,"=C32*$C$42",col="C",bold=True,fill=green); out(44,"=D32*$C$42",col="D",bold=True,fill=green)

section(46,"6.  GROSS CASE VALUE (before risk adjustments)")
label(47,"Gross value (low / high)",bold=True)
out(47,"=$C$16+C44",col="C",bold=True,fill=green); out(47,"=$C$16+D44",col="D",bold=True,fill=green)

section(49,"7.  RISK ADJUSTMENTS")
label(50,"Liability strength (choose)")
lc=ws["C50"]; ws.merge_cells("C50:E50"); lc.value=liab[0][0]; lc.font=input_f; lc.fill=yellow
lc.border=border; lc.alignment=Alignment(indent=1)
label(51,"Liability factor applied"); out(51,"=VLOOKUP($C$50,Reference!$F$2:$G$5,2,FALSE)",PCT)
ws["C51"].alignment=Alignment(horizontal="center")
label(52,"Client comparative fault %",note="Florida: >50% = NO recovery (2023 HB 837)"); inp(52,0,PCT)
label(53,"Comparative-fault status")
fc=ws["C53"]; ws.merge_cells("C53:E53")
fc.value='=IF(C52>0.5,"BARRED - client >50% at fault, no recovery in FL","OK - recovery reduced by fault %")'
fc.font=calc_b; fc.alignment=Alignment(indent=1)
label(54,"Available policy limits (0 = unknown/none)",note="Final value capped here if > 0"); inp(54,0)

section(56,"8.  ADJUSTED CASE VALUE  ->  SETTLEMENT TARGET")
adj_low='=IF($C$52>0.5,0,IF($C$54>0,MIN(C47*$C$51*(1-$C$52),$C$54),C47*$C$51*(1-$C$52)))'
adj_high='=IF($C$52>0.5,0,IF($C$54>0,MIN(D47*$C$51*(1-$C$52),$C$54),D47*$C$51*(1-$C$52)))'
label(57,"SETTLEMENT TARGET RANGE (low / high)",bold=True)
o1=out(57,adj_low,col="C",bold=True,fill=result_fill); o2=out(57,adj_high,col="D",bold=True,fill=result_fill)
o1.font=result_f; o2.font=result_f; ws.row_dimensions[57].height=20
label(58,"Opening demand (target high + 20% cushion)",bold=True,note="Demand full value; don't concede fault in the demand")
out(58,"=D47*1.2",bold=True,fill=yellow)
label(59,"Settlement floor (authority guide)",note="Walk-away guidance; confirm with attorney")
out(59,"=C57")

section(61,"9.  EVALUATE AN INSURANCE OFFER")
label(62,"Insurer's offer"); inp(62,None)
label(63,"Offer vs. target range")
ev=ws["C63"]; ws.merge_cells("C63:E63")
ev.value=('=IF(C62="","- enter offer above -",'
          'IF(C62<C57,"BELOW target - counter high",'
          'IF(C62>D57,"AT/ABOVE target - strong offer","WITHIN target range")))')
ev.font=calc_b; ev.alignment=Alignment(indent=1)
label(64,"Offer as % of target midpoint"); out(64,'=IF(C62="","",C62/((C57+D57)/2))',PCT)
label(65,"Gap to target low / high")
out(65,'=IF(C62="","",C57-C62)',col="C"); out(65,'=IF(C62="","",D57-C62)',col="D")

foot=67; ws.merge_cells(f"B{foot}:E{foot}")
ws[f"B{foot}"]=("Estimate only - not legal advice. Property damage is shown for causation/severity only and is not "
    "part of the recoverable injury value. Valuation depends on venue, witnesses, client presentation, treatment "
    "gaps, prior injuries and policy stacking. Attorney reviews before any demand.")
ws[f"B{foot}"].font=note_f; ws[f"B{foot}"].alignment=Alignment(wrap_text=True,indent=1); ws.row_dimensions[foot].height=40

ws["C39"].comment=Comment("Property damage $ is NOT added to the claim value. Low PD against high medicals signals a causation problem; severe PD supports causation and damages.","Tool")
ws["C52"].comment=Comment("Share of fault attributable to YOUR client (0.20 = 20%). Florida bars recovery above 50%.","Tool")

dv_sev=DataValidation(type="list",formula1="=Reference!$A$2:$A$6",allow_blank=False); ws.add_data_validation(dv_sev); dv_sev.add(ws["C20"])
dv_liab=DataValidation(type="list",formula1="=Reference!$F$2:$F$5",allow_blank=False); ws.add_data_validation(dv_liab); dv_liab.add(ws["C50"])
dv_prior=DataValidation(type="list",formula1="=Reference!$N$2:$N$5",allow_blank=False); ws.add_data_validation(dv_prior); dv_prior.add(ws["C38"])
dv_pd=DataValidation(type="list",formula1="=Reference!$Q$2:$Q$4",allow_blank=False); ws.add_data_validation(dv_pd); dv_pd.add(ws["C40"])

# ---------------- GUIDE ----------------
g=wb.create_sheet("Guide"); g.sheet_view.showGridLines=False
g.column_dimensions["A"].width=2; g.column_dimensions["B"].width=100
def gt(r,t,f):
    c=g[f"B{r}"]; c.value=t; c.font=f; c.alignment=Alignment(wrap_text=True,vertical="top",indent=1)
gt(1,"HOW TO USE THIS TOOL",Font(name=ARIAL,size=14,bold=True,color="1F3864"))
guide=[("Purpose",lbl_b),
 ("Quick, consistent first-pass valuation for preparing demands and evaluating opening offers. A starting point for attorney judgment - not a substitute for it.",lbl_f),
 ("Step by step",lbl_b),
 ("1. Fill the blue cells under ECONOMIC DAMAGES - past/future medicals, wages, out-of-pocket. These are the recoverable specials.",lbl_f),
 ("2. Pick INJURY SEVERITY (yellow dropdown). The multiplier auto-fills; override only with a reason.",lbl_f),
 ("3. Set the PER DIEM rate and number of pain/impact days (injury date to maximum medical improvement is typical).",lbl_f),
 ("4. Enter CASE CHARACTERISTICS - impairment %, age, prior injuries, and property-damage/impact. These adjust pain & suffering only.",lbl_f),
 ("5. Choose LIABILITY STRENGTH, enter the client's COMPARATIVE FAULT %, and policy limits.",lbl_f),
 ("6. Read the SETTLEMENT TARGET RANGE and OPENING DEMAND. Use section 9 to score any offer.",lbl_f),
 ("Methodology",lbl_b),
 ("Multiplier method: non-economic = medical specials x severity multiplier (1.5-10x). Per diem: daily rate x recovery days. The blended low is the lower of the two methods; the high is the higher.",lbl_f),
 ("Property damage - important",lbl_b),
 ("Property damage is NOT a recoverable element of the bodily-injury claim, so it is never added to the case value. It is used only as a causation and severity signal: low property damage against high medical bills flags a causation problem an adjuster will exploit; a totaled vehicle or severe impact makes causation and damages easier to prove and nudges the non-economic value up. The tool runs an automatic causation check comparing property damage to medical specials.",lbl_f),
 ("Other case-characteristic modifiers",lbl_b),
 ("Impairment: factor = 1 + (whole-person % x 1.5), capped 1.6x. Age: younger clients with permanent injuries endure P&S longer, so age brackets scale value. Prior injuries: a prior injury to the SAME body part invites causation/apportionment defenses and reduces value. Impairment, age, prior history and impact severity multiply into one composite modifier (capped 0.4x-2.0x) applied to the non-economic range.",lbl_f),
 ("Risk adjustments",lbl_b),
 ("Gross value x liability factor (1.0 clear -> 0.40 weak), then reduced by comparative fault %, then capped at policy limits if entered.",lbl_f),
 ("Florida law notes",lbl_b),
 ("Modified comparative negligence (Fla. Stat. 768.81, amended 2023 by HB 837): a plaintiff MORE THAN 50% at fault recovers nothing; at/below 50%, damages reduce by the fault %. The tool flags the bar automatically.",lbl_f),
 ("No-fault / PIP for auto: $10,000 PIP covers initial medicals/wages. A bodily-injury claim generally requires meeting the permanent-injury threshold (Fla. Stat. 627.737) before non-economic damages are recoverable - the impairment input helps confirm this.",lbl_f),
 ("No general statutory cap on non-economic damages in ordinary auto/PI cases. Consider UM/UIM and policy stacking when setting limits. The 2023 reforms also shortened the negligence statute of limitations to 2 years (causes accruing on/after 3/24/2023).",lbl_f),
 ("Reference tables",lbl_b),
 ("Severity, liability, age-bracket, prior-history and impact-severity factors all live on the 'Reference' tab. Adjust them to match your firm's playbook; the calculator updates automatically.",lbl_f),
 ("Disclaimer",lbl_b),
 ("Estimates only; not legal advice. Every case turns on its facts, venue, and the people involved. An attorney must review before any demand is sent or offer accepted.",note_f)]
r=3
for t,f in guide:
    gt(r,t,f); g.row_dimensions[r].height=30 if f==lbl_f else (18 if f==lbl_b else 28); r+=1

wb.save("/sessions/zealous-friendly-darwin/mnt/outputs/PI_Case_Valuation_Calculator_FL_tmp.xlsx")
print("saved")
