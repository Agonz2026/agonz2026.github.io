import subprocess, shutil
from openpyxl import load_workbook
def setvals(vals,path="test.xlsx"):
    shutil.copy("PI_Case_Valuation_Calculator_FL.xlsx",path)
    wb=load_workbook(path); ws=wb["Calculator"]
    for k,v in vals.items(): ws[k]=v
    wb.save(path); subprocess.run(["python3","recalc.py",path,"60"],capture_output=True)
    return load_workbook(path,data_only=True)["Calculator"]
def ap(a,b,t=1): return abs(a-b)<=t

# A: full case, low PD -> causation exposure, minor impact factor
ws=setvals({"C11":30000,"C12":10000,"C13":10000,"C15":2000,
 "C20":"3 - Serious (surgery / objective injury)","C27":200,"D27":350,"C28":180,
 "C36":0.10,"C37":35,"C38":"None / clean history","C39":800,
 "C40":"Minor - cosmetic / low impact","C50":"Probable","C52":0.2,"C54":0,"C62":50000})
g=lambda c: ws[c].value
print("econ(no PD)",g("C16"),"medspec",g("C17"))
print("base",g("C32"),g("D32"),"impF",g("D36"),"ageF",g("D37"),"impactF",g("D40"),"comp",round(g("C42"),5))
print("adjNE",round(g("C44")),round(g("D44")),"gross",round(g("C47")),round(g("D47")))
print("target",round(g("C57")),round(g("D57")),"demand",round(g("C58")))
print("causation:",g("C41"))
assert ap(g("C16"),52000), "PD must NOT be in economic total"
assert ap(g("C17"),40000) and ap(g("C32"),36000) and ap(g("D32"),180000)
assert ap(g("D36"),1.15) and ap(g("D37"),1.05,0.001) and ap(g("D40"),0.85)
assert ap(g("C42"),1.026375,0.001)
assert ap(g("C44"),36949.5,1) and ap(g("D44"),184747.5,1)
assert ap(g("C47"),88949.5,1) and ap(g("D47"),236747.5,1)
assert ap(g("C57"),60485.66,1) and ap(g("D57"),160988.3,1)
assert ap(g("C58"),284097,1)
assert "causation EXPOSURE" in g("C41")
print("A PASS\n")

# B: severe impact -> supportive causation message + 1.10 factor
ws=setvals({"C11":50000,"C20":"4 - Severe (permanent / disfigurement)","C28":0,
 "C36":0.30,"C37":25,"C38":"Prior - same body part (documented)","C39":40000,
 "C40":"Severe - totaled / major impact","C50":"Clear / strong","C52":0,"C54":0})
print("impactF",ws["D40"].value,"comp",round(ws["C42"].value,5),"causation:",ws["C41"].value)
# impF=1.45, ageF=1.15, priorF=0.80, impactF=1.10 -> 1.45*1.15*0.8*1.1=1.46740
assert ap(ws["D40"].value,1.10) and ap(ws["C42"].value,1.4674,0.001)
assert "well-supported" in ws["C41"].value
assert "likely met" in ws["C43"].value
print("B PASS\n")

# C: PD present & consistent (not low, not severe)
ws=setvals({"C11":20000,"C20":"2 - Moderate (treatment, minor residual)","C28":0,
 "C36":0,"C37":45,"C39":15000,"C40":"Moderate - drivable with damage","C52":0})
print("causation:",ws["C41"].value)
assert "consistent" in ws["C41"].value
print("C PASS\n")

# D: >50% fault barred
ws=setvals({"C11":50000,"C20":"2 - Moderate (treatment, minor residual)","C28":0,"C52":0.6})
assert ws["C57"].value==0 and ws["D57"].value==0 and "BARRED" in ws["C53"].value
print("D PASS (>50% bar)\n")

# E: policy cap
ws=setvals({"C11":200000,"C12":100000,"C20":"5 - Catastrophic (disability, brain/spinal)",
 "C27":500,"D27":1000,"C28":730,"C36":0.40,"C37":20,"C40":"Severe - totaled / major impact",
 "C50":"Clear / strong","C52":0,"C54":250000})
print("comp",round(ws["C42"].value,4),"target high (cap 250k)",ws["D57"].value)
assert ws["D57"].value==250000 and ws["C57"].value==250000
print("E PASS (policy cap)")
print("ALL CHECKS PASSED")
